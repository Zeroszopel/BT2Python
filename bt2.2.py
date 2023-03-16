import openpyxl
from Student import Student

listSV = []
mini = [3, 10]
gioi = 0
kha = 0
tb = 0


def getlistSV():
    return listSV


def SortCri(e):
    return e.MaSV


def tinhdtb(a, b, c):
    return round((a + b + c) / 3, 1)


def checkHK(e):
    if e >= 8:
        return "Giỏi"
    elif e >= 6.5:
        return "Khá"
    else:
        return "Trung Bình"


def max(i):
    max = 0
    if i == 0:
        for j in range(len(listSV)):
            if max < len(listSV[j].MaSV):
                max = len(listSV[j].MaSV)

    if i == 1:
        for j in range(len(listSV)):
            if max < len(listSV[j].Ho + " " + listSV[j].Ten):
                max = len(listSV[j].Ho + " " + listSV[j].Ten)

    if max < mini[i]:
        max = mini[i]

    return max


def readexcel():
    global hk
    global filename
    hk = [0, 0, 0]
    filename= input("Nhập tên file cần đọc:")+'.xlsx'
    book = openpyxl.load_workbook(filename)
    sh = book.active
    maxrow = sh.max_row
    i = 1
    t = "A" + str(i)
    while sh[t].value != "STT" or i == maxrow:
        i += 1
        t = "A" + str(i)
    if i == maxrow:
        print("Không đọc dược dữ liệu")
    else:
        global index
        index = i
        j = i + 1
        t = "A" + str(j)
        while type(sh[t].value) == int:
            stu = Student(sh.cell(row=j, column=2).value,
                          sh.cell(row=j, column=3).value,
                          sh.cell(row=j, column=4).value,
                          sh.cell(row=j, column=5).value,
                          round(float(sh.cell(row=j, column=6).value), 1),
                          round(float(sh.cell(row=j, column=7).value), 1),
                          round(float(sh.cell(row=j, column=8).value), 1),
                          0, ""
                          )
            stu.DTB = tinhdtb(stu.Toan, stu.Ly, stu.Hoa)
            stu.HK = checkHK(stu.DTB)
            if stu.HK == "Giỏi":
                hk[0] += 1
            elif stu.HK == "Khá":
                hk[1] += 1
            else:
                hk[2] += 1

            listSV.append(stu)
            j += 1
            t = "A" + str(j)
        listSV.sort(key=SortCri)


def showsv(listSV):
    totallen = 0
    for i in range(2):
        totallen += max(i)

    print("_" * (totallen + 39))
    print(
        "| {} | {} | {:^5} | {:^5} | {:^5} | {:^5} |".format("STT".center(max(0)), "Họ và Tên".center(max(1)),
                                                             "Toán", "Lý",
                                                             "Hóa", "ĐTB", "HK"))
    for i in listSV:
        print(
            "| {:} | {:} | {:^5} | {:^5} | {:^5} | {:^5} |".format(i.MaSV.center(max(0)),
                                                                   (i.Ho + " " + i.Ten).center(max(1)),
                                                                   i.Toan, i.Ly, i.Hoa,
                                                                   i.DTB, i.HK))
    print("|" + "_" * (totallen + 37) + "|")


def inexcel(list):
    book = openpyxl.load_workbook(filename)
    sh = book.active
    for i in range(1, len(list) + 1):
        sh.cell(row=(i + index), column=2).value = list[i - 1].MaSV
        sh.cell(row=(i + index), column=3).value = list[i - 1].Ho
        sh.cell(row=(i + index), column=4).value = list[i - 1].Ten
        sh.cell(row=(i + index), column=5).value = list[i - 1].NgaySinh
        sh.cell(row=(i + index), column=6).value = list[i - 1].Toan
        sh.cell(row=(i + index), column=7).value = list[i - 1].Ly
        sh.cell(row=(i + index), column=8).value = list[i - 1].Hoa
        sh.cell(row=(i + index), column=20).value = list[i - 1].DTB
    sh.cell(row=(index + len(list) + 1), column=18).value = "Giỏi"
    sh.cell(row=(index + len(list) + 2), column=18).value = hk[0]
    sh.cell(row=(index + len(list) + 1), column=19).value = "Khá"
    sh.cell(row=(index + len(list) + 2), column=19).value = hk[1]
    sh.cell(row=(index + len(list) + 1), column=20).value = "Trung Bình"
    sh.cell(row=(index + len(list) + 2), column=20).value = hk[2]
    book.save(filename)
    print("Đã cập nhật file excel: "+filename)

check = -1
while check != 0:
    print("1. Nhập dữ liệu sinh viên")
    print("2. Xem danh sách sinh viên")
    print("3. In ra file excel")
    check = int(input("Nhập số: "))
    while check < 0 or check > 3:
        check = int(input("Nhập lại số: "))

    print("*" * 50)
    if check == 1:
        readexcel()
    elif check == 2:
        showsv(getlistSV())
    elif check == 3:
        inexcel(getlistSV())
    else:
        break

    print("*" * 50)
